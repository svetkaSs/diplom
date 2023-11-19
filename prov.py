import sys
from fer import FER
import cv2
from moviepy.editor import VideoFileClip
import os
import openpyxl as ox
import matplotlib.pyplot as plt
from moviepy.editor import * #подключаем пакет moviepy
sys.path.append('./esrApp_tst')
sys.path.append('./esrApp_tst/classificator')
sys.path.append('./esrApp_tst/level-generation')
from esrApp_tst.mainEsr import AudioRecognition


def save_frame_range_sec(video_path, start_sec, stop_sec, step_sec, dir_path, basename, ext='jpg'):
    cap = cv2.VideoCapture(video_path)


    if not cap.isOpened():
        return


    os.makedirs(dir_path, exist_ok=True)
    base_path = os.path.join(dir_path, basename)
    emot = list()
    emot_ver=list()
    cols=['A','B','C']
    wb = ox.Workbook()
    ws = wb.worksheets[0]
    i=1


    digit = len(str(int(cap.get(cv2.CAP_PROP_FRAME_COUNT))))

    fps = cap.get(cv2.CAP_PROP_FPS)
    fps_inv = 1 / fps
    # MyFile = open('output1.txt', 'w')
    sec = start_sec
    while sec < stop_sec:
        n = round(fps * sec)
        cap.set(cv2.CAP_PROP_POS_FRAMES, n)
        ret, frame = cap.read()
        if ret:

            name = './data/frame_angry_m' + str(sec) + '.jpg'
            #print('Creating...' + name)
            cv2.imwrite(name, frame)
            test_img = cv2.imread(name)
            plt.imshow(test_img[:, :, ::-1])
            emo_detector = FER(mtcnn=True)
            captured_emotions = emo_detector.detect_emotions(test_img)
            print(captured_emotions)

            dominant_emmotion, emotion_score = emo_detector.top_emotion(test_img)
            emot.append(dominant_emmotion)
            emot.append(emotion_score)

            #print(dominant_emmotion, emotion_score)
            #print( sec)
            sec += step_sec

            # MyFile.write(dominant_emmotion)
            # MyFile.write(' ')
            # MyFile.write(str(emotion_score))
            # MyFile.write('\n')
            # ws.cell(row=i, column=1).value = dominant_emmotion
            # ws.cell(row=i, column=2).value = emotion_score
            # ws.cell(row=i, column=3).value=sec
            # i +=1
            # wb.save('./Alg1_01-01-08-02-01-01-21.xlsx')


        cv2.imwrite(
                '{}_{}_{:.2f}.{}'.format(
                    base_path, str(n).zfill(digit), n * fps_inv, ext
                ),
                frame

            )

video_path="/home/karakul/diploma/diplom/video-files/01-01-03-02-01-01-13.mp4"

audioclip = AudioFileClip(video_path) #видеофайл 1.mp4
print (audioclip)

audioclip.write_audiofile("./output-audio/out_audio.wav") #извлеченная аудиодорожка в файл out_audio.mp3

audio_recognition =AudioRecognition('./output-audio/out_audio.wav')
audio_recognition.getAudioRecognition()
clip = VideoFileClip(video_path)
save_frame_range_sec(video_path,0, clip.duration, 0.03,
                     './data', 'sample_video_img')





